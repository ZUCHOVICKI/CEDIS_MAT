

import openpyxl
from pathlib import Path
from datetime import datetime
from openpyxl.styles import Color, PatternFill, Font, Border
from OpenOrders import *

##CENTRAL FILE WHERE OPERATIONS AND DATA IS STORED
xlsx_file = Path('.', 'CEDIS_MAT.xlsx')

wb_obj = openpyxl.load_workbook(xlsx_file)

## mAIN FUCNTION FOR ADDING ORDERS TO AND COMPONENTS ( DATE = "DATE WHEN THE ORDER IS DUE" 
#  DATA = ITEMS THAT MUST BE LOADED TO THE ORDER FILE) 
# ++IMPORTANT ADDORDER MUST BE CALLED AFTER EVERY NEW COMPONENT IS ADDED TO THE ORDER
def ADDORDER(Data,Date):


    ## USED TO LOEADING NEW COMPONENTS ON EXISTING ORDERS WHITHOUT DELETING EXISITING COMPONENTS
    try:
        dataDict =  finishedOrders[Date] 
    except:
        dataDict = {}

    ## GETS THE EXCEL SHEET WHERE INFORMATIÓN IS LOCATED AND THE NUMBER OF TIMES IT NEEDS TO BE CALLED 
    for key in Data:
        sheetnom = key
        num = Data.get(key)
        
        
    ##SEARCHS FOR THE WORKSHEET AND TAKES ALL COMPONENTS NEEDED FOR THE ORDER 
    for sheet in wb_obj.worksheets:
        
        if sheet.title != sheetnom:
            continue
        
        for row in sheet.iter_rows(max_row=sheet.max_row):

            if row[1].value == None:
                continue
            if row[3].value == None:
                continue

            if row[1].value in dataDict:
                
                dataDict[row[1].value] += (row[3].value)*num

            else:
                dataDict[row[1].value] = (row[3].value)*num

    ##ADDORDER LOADS DATA ON THE GLOBAL VARIABLE finishedOrders
    finishedOrders[Date] = dataDict
    
##GETS ALL THE INVENTAROY COMPONETS AND THE NUMBER ONHAND FROM A SPECIFIC EXCEL PAGE
##**TODO AGREGAR 
def GetInventario():
    inventario = {}

    onHand = wb_obj["ON-HAND"]
    HojasInventario =[onHand]



    for sheet in HojasInventario:
        
        for row in sheet.iter_rows(max_row=sheet.max_row):
                
                if row[0].value == None:
                    continue
                if row[1].value == None:
                    continue

                if row[0].value in inventario:
                    
                    inventario[row[0].value] += (row[1].value)

                else:
                    inventario[row[0].value] = (row[1].value)
    ##GETINVENTARIO LOADS THE AVAILABLE MATERIALS ON GLOBAL VARIABLE inventario
    return inventario

def GetInventarioFull():
    inventarioFull = {}

    onHandFull = wb_obj["ON-HAND-Full"]
    HojasInventario =[onHandFull]



    for sheet in HojasInventario:
        
        for row in sheet.iter_rows(max_row=sheet.max_row):
                
                if row[0].value == None:
                    continue
                if row[1].value == None:
                    continue

                if row[0].value in inventarioFull:
                    
                    inventarioFull[row[0].value] += (row[1].value)

                else:
                    inventarioFull[row[0].value] = (row[1].value)
    ##GETINVENTARIO LOADS THE AVAILABLE MATERIALS ON GLOBAL VARIABLE inventario
    return inventarioFull


## CHECKS FOR THE TOTAL NUMBER OF MATERIALS AFTER THE ORDER HAS BEEN FINALIZED ( Orders = finished orders and date
#  Inventary = inventory details order = dates ordered from closest to furthest)
def CheckAvailable(Orders,Inventary,order):
    Actualdate = ""
    warningsProduct = {}
    
    for date in order:
       Actualdate =date
       
       data =  Orders[date]
      
      
       for product in data:
           productData = product
           if product == 301303:
                product = str(product)
           if product not in Inventary:
              
               if product not in warningsProduct:
                
                warningsProduct[product]= 0 - data[productData]
               else:
                warningsProduct[product] -= (data[productData])
               continue
           
           Inventary[product] -= data[productData]
        #    inventarioFull[product] -= data[product]
           
           
           
           if(Inventary[product]<=0):
               
               
               if product not in warningsProduct:
                
                warningsProduct[product] = (Inventary[product] )

            
           
       newOrders = warningsProduct.copy()
       
       
       dateWarnings[Actualdate]=newOrders
       newOrders= {}
    #    print("Guardado")
    
    return dateWarnings


 ##Function for display of all loaded Bills of Materials on the interfaze   
def getAllBills():
    MATBills = {}
    count = 1
    for sheet in wb_obj.worksheets:
        if sheet.title in ["ON-HAND","Ordenes"]:
            continue
        MATBills[count] = sheet.title
        count+=1  
    count = 1     
    # for x in MATBills:
    #             print("{0} ---- {1}".format(count,MATBills.get(x)))
    #             count+=1
    return MATBills





    ##Function for ordering dates after the creatión of an order      
def OrderDate(date):
    if date not in dateOrder:
     dateOrder.append(date)
    str(date)
    dateOrder.sort(key=lambda date: datetime.strptime(date, "%d/%m/%Y"))
    # print(finishedOrders)
    
##Function for adding onlly one subcomponent from a bill and not the whole bill
def AgregarIndividual(Item,Date,Amount):
   

    
    try:
        finishedOrders[Date][Item] += Amount
    except:
        finishedOrders[Date] = {Item:Amount}
    

##Final function that generates an Excel file using the Availability of the materials
def GenerarReporte(Warnings,FirstDate,Orders,DateOrder,inventario,HR=0):
    
    items =  ItemMaster()
    today = datetime.today().strftime('%Y-%m-%d')
    wb_obj.create_sheet(title="Reporte")
    wb_obj.create_sheet(title="ReporteManufactura")
    sheet = wb_obj["Reporte"]
    sheetManufactura = wb_obj["ReporteManufactura"]

    redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')
    greenfill = PatternFill(start_color='00FF00',
                   end_color='00FF00',
                   fill_type='solid')
    openOrders = GetOpenOrders()


    for s in wb_obj.worksheets:

        
        if s.title != 'Reporte' and s.title != 'ReporteManufactura':
            # print(s.title)
            sheet_name = wb_obj[s.title]
            wb_obj.remove_sheet(sheet_name)

    ##REPORTE DE P
    sheet.cell(column=1, row=1, value="Fecha de Generación de Reporte")
    sheet.cell(column=2, row=1, value=today)

    ##REPORTE DE M
    sheetManufactura.cell(column=1, row=1, value="Fecha de Generación de Reporte")
    sheetManufactura.cell(column=2, row=1, value=today)


    sheet.cell(column=10, row=1, value="Fecha Maxima de Materiales")
    sheet.cell(column=11, row=1, value=FirstDate)

    currentRow = 3
    # print(Warnings)
    for Date in DateOrder:

        sheet.cell(column=1, row=currentRow, value="Orden")
        sheet.cell(column=2, row=currentRow, value=Date)
        currentRow+=1
        sheet.cell(column=1, row=currentRow, value="Material")
        sheet.cell(column=2, row=currentRow, value="# Necesitado")        
        sheet.cell(column=3, row=currentRow, value="Total")
        sheet.cell(column=4, row=currentRow, value="Inventario Total")
        sheet.cell(column=5, row=currentRow, value="OpenOrder")
        sheet.cell(column=6, row=currentRow, value="Status")

        ##OPEN ORDER + INVENTARIO ACTUAL 
        ##TRANSITO

        currentRow+=1
        for item in Orders[Date]:
            itemStr = str(item)
            # print(itemStr)
            # print(type(itemStr))

            
            try:
                if items[itemStr] == "M":
                    # print("Manufactura")
                    continue
                
            except:
                
                
                print("Not In Item Master")

            
            sheet.cell(column=1, row=currentRow, value=item)
            sheet.cell(column=2, row=currentRow, value=Orders[Date][item])
            if item in Warnings[Date]:
                sheet.cell(column=3, row=currentRow, value=Warnings[Date][item])
                
                sheet.cell(column=6, row=currentRow).fill=redFill
            else:
                sheet.cell(column=3, row=currentRow, value=inventario[itemStr])
                sheet.cell(column=6, row=currentRow).fill=greenfill

            item = str(item)
            try:
                sheet.cell(column=4, row=currentRow, value=inventarioFull[item])
            except:
                sheet.cell(column=4, row=currentRow, value=0)
            currentRow+=1


            if item in openOrders:
                sheet.cell(column=5, row=currentRow, value=openOrders[item])

    currentRow = 3    
    for Date in DateOrder:

        sheetManufactura.cell(column=1, row=currentRow, value="Orden")
        sheetManufactura.cell(column=2, row=currentRow, value=Date)
        currentRow+=1
        sheetManufactura.cell(column=1, row=currentRow, value="Material")
        sheetManufactura.cell(column=2, row=currentRow, value="# Necesitado")        


    

    ##OPEN ORDER + INVENTARIO ACTUAL 
    ##TRANSITO

        currentRow+=1
        for item in Orders[Date]:
            itemStr = str(item)
           

            
            try:
                if items[itemStr] == "P":
                    # print("Manufactura")
                    continue
                
            except:
                
                
                print("Not In Item Master")

            
            sheetManufactura.cell(column=1, row=currentRow, value=item)
            sheetManufactura.cell(column=2, row=currentRow, value=Orders[Date][item])
           
            currentRow+=1
# report is saved with the keyword "Reporte" + the date it was generated


    if HR == 0:
     wb_obj.save('./Reportes/Reporte--'+today+'.xlsx')
    else:
     wb_obj.save('./Reportes/HighRunners--'+today+'.xlsx')

##Loads a Central DB where information about all materials is stored
def ItemMaster():
    ItemS = {}
    xlsx_file = Path('.', 'Item master.xlsx')
    Item_Master = openpyxl.load_workbook(xlsx_file)

    sheet = Item_Master["Data"]
    for row in sheet.iter_rows(max_row=sheet.max_row):
            
            ItemS[row[1].value] = row[10].value
    
    return ItemS

## loads all the components in a selected Bill of materials and siplays it on the interface
def GetIndividualComponents(Item):
    allItem = []
    for sheet in wb_obj.worksheets:
        
        if sheet.title != Item:
            continue
        
        for row in sheet.iter_rows(max_row=sheet.max_row):
            if row[1].value == None:
                continue
            if row[3].value == None:
                continue
            allItem.append(row[1].value)
        
    return allItem
MATBills = getAllBills()


def DeleteALL():
    finishedOrders.clear() ##Orders that have been submitted
    dateOrder.clear() # Orders submitted but in chronological order
    dateWarnings.clear() # Orders that have one or more materials close or under 0 available units
    ActualOrder.clear() #Private Variable used in functions for temporary storage of order info
    dataDict.clear()


def DeleteOne(Date):
    finishedOrders.pop(Date,None)##Orders that have been submitted

    try:
        dateOrder.remove(Date)
    except:
        print("***ERROR NOT IN DATE ARRAY SYSTEM***")
    dateWarnings.pop(Date,None) # Orders that have one or more materials close or under 0 available units
    ActualOrder.pop(Date,None) #Private Variable used in functions for temporary storage of order info
    dataDict.pop(Date,None) 


def GetHighRunners():

    FinalRunners = {}
    content = {}
    contentMGR = []
    TitleMGR = []
    RunnerTitle = ""
    HighRun = wb_obj["HighRunner"]

    for row in HighRun.iter_rows(max_row=HighRun.max_row):
        if row[0].value == "Level":
            continue
        if row[0].value == "Parent":
            
            RunnerTitle = row[1].value
            
            content.clear()
            continue

        content[row[1].value] = row[3].value
        FinalRunners[RunnerTitle] = content.copy() 
     
   
    return FinalRunners







## Global variables that can be read from all the aplication 
finishedOrders = {} ##Orders that have been submitted
dateOrder = [] # Orders submitted but in chronological order
dateWarnings = {} # Orders that have one or more materials close or under 0 available units
ActualOrder = {} #Private Variable used in functions for temporary storage of order info
dataDict = {} #Private Variable used in functions for temporary storage of componets info
inventario = GetInventario() # Current inventory loaded on separate file

inventarioFull = GetInventarioFull()#All inventory loaded on separate file

Runners = GetHighRunners() ## HIGH RUNNERS / IMportant structures


