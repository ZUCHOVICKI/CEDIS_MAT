##Finds the file with all inventory and downloads it to a "ON-HAND" sheet on central Excel

import openpyxl
from pathlib import Path
import os
xlsx_file = Path('.', 'CEDIS_MAT.xlsx')

wb_obj = openpyxl.load_workbook(xlsx_file)



try:
    full = wb_obj['ON-HAND-Full']
    wb_obj.remove(full)
    wb_obj.create_sheet(title="ON-HAND-Full")
except:
    wb_obj.create_sheet('ON-HAND-Full')


full = wb_obj['ON-HAND-Full']
inv = wb_obj["ON-HAND"]

inventario ={}
inventarioFull = {}

path = "./inventario"

Folderlenght = len(os.listdir(path))
def AddInventario(SelectedLocation,SelectedSite):
    inv = wb_obj["ON-HAND"]
    for i in range(0,Folderlenght):
        path = "./inventario"
        

        if len(os.listdir(path)) == 0:
            print("No hay Archivos en la Carpeta de inventario")
            raise SystemExit(0)

        finalFile = os.listdir(path)[0]

        print(finalFile)
        f= open(path+"/"+finalFile, "r") 

        valores = []
        
        SelectedSite = list(set(SelectedSite)) 
        print(SelectedSite)
        for x in f:
            valores = []
            for Site in SelectedSite:
                if x.startswith(Site):
                    valores = x.split()
                    
                    if valores[1] not in SelectedLocation:
                        
                        continue
                    if len(valores)>=12:
                        valores.pop(3)
                    
                
                    if valores[2] in inventario:
                        inventario[valores[2]] += float(valores[4].replace(',', ''))
                    else:
                        inventario[valores[2]] = float(valores[4].replace(',', ''))
            #  print(valores)
                    print(inventario)
                else:
                    continue
    
            valores = []
            valores = x.split()
            
            if  valores==[]:
                # print(valores)
                continue
            if valores[0] in excludedData:
                # print(valores)
                continue
            if len(valores)>=12:
                valores.pop(3)
            if valores[2] in inventarioFull:
                inventarioFull[valores[2]] += float(valores[4].replace(',', ''))
            else:
                inventarioFull[valores[2]] = float(valores[4].replace(',', ''))       
            
        f.close() 
        

    print(inventario)




        
    wb_obj.remove(inv)
    wb_obj.create_sheet(title="ON-HAND")
    inv = wb_obj["ON-HAND"]

    row = 1

    for x in inventario:
        inv.cell(column=1, row=row, value=x)
        inv.cell(column=2, row=row, value=inventario[x])
        row +=1

    row = 1
    for x in inventarioFull:
        full.cell(column=1, row=row, value=x)
        full.cell(column=2, row=row, value=inventarioFull[x])
        row +=1

    wb_obj.save(filename = "CEDIS_MAT.xlsx")

availableSites = []

excludedData = ['iclorp.p', 'Page:', 'Lot/Serial', 'Site', '--------', 'End', 'Report', 'Item', 'Site:', 'Location:', 'Output:', 'Batch']

def GetLocation():
      for i in range(0,Folderlenght):
        path = "./inventario"
        

        if len(os.listdir(path)) == 0:
            print("No hay Archivos en la Carpeta de inventario")
            raise SystemExit(0)

        finalFile = os.listdir(path)[0]

        # print(finalFile)
        f= open(path+"/"+finalFile, "r") 

        valores = []
        
        for x in f:
            valores = []
           
            valores = x.split()
            if valores == []:
                continue
                   
            if valores[0] == None:
                continue
            else:
                # print(valores[0])
                if valores[0] in availableSites or valores[0] in excludedData:
                    continue
                else:
                    availableSites.append(valores[0])
    
        f.close() 
        # os.remove(os.path.join(path, finalFile))
availableLocations = []
def SiteLocations(site):
    for i in range(0,Folderlenght):
        path = "./inventario"
        

        if len(os.listdir(path)) == 0:
            print("No hay Archivos en la Carpeta de inventario")
            raise SystemExit(0)

        finalFile = os.listdir(path)[0]

        # print(finalFile)
        f= open(path+"/"+finalFile, "r") 

        valores = []
        
        for x in f:
             valores = []
           
             valores = x.split()
             if valores == []:
                    continue
                   
             if valores[0] == None:
                 continue
             else:
                    # print(site)
                    # print(valores)
                    SelectedSites.append(site)
                    if valores[0]==site:
                        if valores[0] in excludedData:
                            continue
                        if valores[1] in availableLocations:
                            continue
                        else:
                            availableLocations.append(valores[1])
    
        f.close() 





SelectedLocation = []
SelectedSites= []




# GetLocation()
# site = availableSites[0]
# SiteLocations(site)
# print(availableSites)
# print(availableLocations)

# SelectedLocation.append(availableLocations[0])
# SelectedLocation.append(availableLocations[-1])

# print(SelectedLocation)

# AddInventario(SelectedLocation,SelectedSites)
## POSICIÓN / VALOR 
# 1 = Site 
# 2 =Location 
# 3 =Item Number 
# 4 =UM
# 5 = Qty on Hand
# 6 = Created
# 7 = Expire
# 8 =    Assay % Grade 
# 9 = Status     
# 10 = Avail
# 11 = Net
# 12 = OvrIs