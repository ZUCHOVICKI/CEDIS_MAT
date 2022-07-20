import openpyxl
from pathlib import Path
import os


def GetOpenOrders():

    path = "./OpenOrder"

    Folderlenght = len(os.listdir(path))
    count = 0

        
    for i in range(0,Folderlenght):

        Openorders = {}

       

        title = ""
        if len(os.listdir(path)) == 0:
            print("No hay Archivos en la Carpeta de inventario")
            raise SystemExit(0)

        finalFile = os.listdir(path)[count]

        count+=1

        print(finalFile)

        xlsx_file = Path(path, finalFile)
    
        wb_obj = openpyxl.load_workbook(xlsx_file)

        sheet = wb_obj.active

        for row in sheet.iter_rows(max_row=sheet.max_row):


            if "PRO" in row[4].value:
                print(row[4].value + " " + str(row[9].value))

                if row[8].value in Openorders:
                    Openorders[row[8].value]  += row[9].value
                else:
                    Openorders[row[8].value]  = row[9].value

                

      

        wb_obj.close()

        return Openorders

