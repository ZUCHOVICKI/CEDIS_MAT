import openpyxl
import numpy as np
from pathlib import Path
from datetime import date, datetime , timedelta
from openpyxl.styles import Color, PatternFill, Font, Border
import os

def last_day(year, month):
    last_date = datetime(year, month + 1, 1) + timedelta(days=-1)
    return last_date.strftime("%Y-%m-%d")


xlsx_file = Path('.', 'RunTime.xlsx')

wb_obj = openpyxl.load_workbook(xlsx_file)

sheet = wb_obj["Data"]

RunTimesMat = {}
RunTimeExternal = {}

TimeFHO = 96
Time9GURUEM = 168

for row in sheet.iter_rows(max_row=sheet.max_row):
    if row[0].value =="Routing Code":
        continue
    if row[5].value == "FHO":
        RunTimeExternal[row[0].value]= "PROVEDOR EXTERNO FHO"
        continue
    if row[5].value == "9GURUEM":
        RunTimeExternal[row[0].value]= "PROVEDOR EXTERNO 9GURUEM"
        continue
    if row[0].value in RunTimesMat:
        # print(row[0].value)
        # print(RunTimesMat[row[0].value])
        RunTimesMat[row[0].value] += float(row[8].value)
    else:
        RunTimesMat[row[0].value]= float(row[8].value)




path = "./Reportes"

Folderlenght = len(os.listdir(path))
count = 0

    
for i in range(0,Folderlenght):

    orders = {}

    materials = {}

    title = ""
    if len(os.listdir(path)) == 0:
        print("No hay Archivos en la Carpeta de inventario")
        raise SystemExit(0)

    finalFile = os.listdir(path)[count]
    
    print(finalFile)

    xlsx_file = Path(path, finalFile)

    wb_obj = openpyxl.load_workbook(xlsx_file)

    sheet = wb_obj["ReporteManufactura"]

    

    for row in sheet.iter_rows(max_row=sheet.max_row):

        ExcludedData = ["Fecha de Generación de Reporte","Material"]

        if row[0].value == "Orden":
            title = row[1].value
            
            materials.clear()
            continue

        if row[0].value in ExcludedData:
            continue
        if row[0].value not in ExcludedData or row[0].value != "":

            materials[str(row[0].value)] = row[1].value
            # print(materials)
            orders[title] = materials.copy()
            
        
    orders.pop('')
    OrdersByTime = {}
    Time = 0
    TimeOffFactoryFHO = 0
    TimeOffFactory9GURUEM = 0
    countFHO = 0
    count9GURUEM = 0

    AllTime = 0
    AllOffFactory = 0
    # print("Orders")
    # print(orders)
    # print("Times")
    # print(RunTimesMat)
    for x in orders:
        
        AllTime = Time
        Time = 0
        for y in orders[x]:
        
            try:
                Time += RunTimesMat[y]*orders[x][y]
            except:
                continue

       

        OrdersByTime[x] = {"Time" : Time}
    count +=1
    print(OrdersByTime)


    wb_obj.create_sheet("Reporte_Tiempo")

    reporte = wb_obj["Reporte_Tiempo"]
    today = date.today()
    todayExcel = datetime.today().strftime('%Y-%m-%d')
    row = 0
   
    Todaymonth = datetime.now().month
    Todayyear = datetime.now().year
    Todayday = datetime.now().day
    endMonth = last_day(Todayyear,Todaymonth)
    days = np.busday_count( today, endMonth )
    workinghours = days*9.5
    ExternalHours = days*24
    HoursCount = {"Plant":workinghours,"OffPlant":ExternalHours}


    redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')

    greenfill = PatternFill(start_color='00FF00',
                   end_color='00FF00',
                   fill_type='solid')


    reporte.cell(column=1, row=1, value="Fecha de Generación de Reporte")
    reporte.cell(column=2, row=1, value=todayExcel)

    reporte.cell(column=10, row=1, value="Fecha Actual")
    reporte.cell(column=11, row=1, value=todayExcel)
    

    reporte.cell(column=10, row=2, value="Fecha Fin de Mes")
    reporte.cell(column=11, row=2, value=endMonth)

    reporte.cell(column=10, row=2, value="Horas Disponibles")
    reporte.cell(column=11, row=2, value=workinghours)
    
    reporte.cell(column=10, row=3, value="Horas Proovedor disponibles")
    reporte.cell(column=11, row=3, value=ExternalHours)

    reporte.cell(column=10, row=5, value="Horas Internas Disponibles")
    reporte.cell(column=11, row=5, value="Status")
    reporte.cell(column=12, row=5, value="Horas Externas Disponibles")


    row = 3


    
    for x in OrdersByTime:
        row+=1
        reporte.cell(column=1, row=row, value="Fecha de orden")
        reporte.cell(column=2, row=row, value=x)
        row+=1
        reporte.cell(column=1, row=row, value="Número de Parte")
        reporte.cell(column=2, row=row, value="Cantidad de Numeros de Parte")
        reporte.cell(column=3, row=row, value="Horas de Manufactura Interna")
        reporte.cell(column=4, row=row, value="FHO")
        reporte.cell(column=5, row=row, value="9GURUEM")
        row+=1
        for y in orders[x]:
            try:
                reporte.cell(column=1, row=row, value=y)
                reporte.cell(column=2, row=row, value=orders[x][y])
                reporte.cell(column=3, row=row, value=(orders[x][y])*RunTimesMat[y])
                HoursCount["Plant"] -= orders[x][y]*RunTimesMat[y]

                
                reporte.cell(column=11, row=row, value= HoursCount["Plant"])
                
                if HoursCount["Plant"] <0:
                    reporte.cell(column=12, row=row).fill = redFill
                else:
                    reporte.cell(column=12, row=row).fill = greenfill
                row+=1
                
            except:
                reporte.cell(column=1, row=row, value=y)
                if RunTimeExternal[y] =="PROVEDOR EXTERNO FHO":
                   
                    
                        reporte.cell(column=4, row=row, value="X")
                        if orders[x][y] >= 15:
                            countFHO += 1 

                if RunTimeExternal[y] =="PROVEDOR EXTERNO 9GURUEM":
                    
                    
                        reporte.cell(column=5, row=row, value="X")
                        if orders[x][y] >= 15:
                            count9GURUEM += 1 

                
      
                row+=1
   

        print("Count FHO " + str(countFHO))
        print("Count 9Guruem " + str(count9GURUEM))

        HoursCount["OffPlant"] -= TimeFHO+(countFHO*24)
        print("TIME FHO "+str(HoursCount["OffPlant"]))

        HoursCount["OffPlant"] -= Time9GURUEM+(count9GURUEM*24)
        print("TIME 9Guruem "+str(HoursCount["OffPlant"]))
        
        reporte.cell(column=12, row=row, value= HoursCount["OffPlant"])
        if HoursCount["OffPlant"]<0 :
                reporte.cell(column=13, row=row).fill = redFill
        else:
            reporte.cell(column=13, row=row).fill = greenfill
            
    wb_obj.save('./Manufactura/ReporteManufactura--'+finalFile+'.xlsx')




