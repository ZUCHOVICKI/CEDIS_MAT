import openpyxl
from pathlib import Path
import os
xlsx_file = Path('.', 'CEDIS_MAT.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file)


##Global variables accesed by whole application
bill = {}
billOrder = {}
billDesc = {}
Billtitle = ""
values = []


##Dynamically findd all the files in the folder and create a new sheet with Bill of material's data
path = "./Bills"

Folderlenght = len(os.listdir(path))

for i in range(0,Folderlenght):
    path = "./Bills"
    

    if len(os.listdir(path)) == 0:
        print("No hay Archivos en la Carpeta de inventario")
        raise SystemExit(0)

    finalFile = os.listdir(path)[0]

    print(finalFile)
    f= open(path+"/"+finalFile, "r") 
##function reads each row and tajes the data neccesary for creating the Excel structure
    for x in f:
        if x.startswith("Parent") or x.startswith(".") or x[:0].isdigit() or x[:1].isdigit() :
            values=[]
            values = x.split()
            
            if values[0] == "Parent":
                Billtitle = values[1]
                continue
            if values[1] in bill:
                bill[values[1]] += float(values[-2].replace(',', ''))
                billOrder[values[1]] = values[0]
                billDesc[values[1]] = values[2]
            else:
                bill[values[1]] = float(values[-2].replace(',', ''))
                billOrder[values[1]] = values[0]
                billDesc[values[1]] = values[2]
    


    
    f.close() 
    os.remove(os.path.join(path, finalFile))

##Creation of new sheet and deletion of pre-existing one
if Billtitle in wb_obj.sheetnames:
    sheet = wb_obj[Billtitle]
    wb_obj.remove(sheet)
    
wb_obj.create_sheet(title=Billtitle)
sheet = wb_obj[Billtitle]

row = 1

for x in bill:
    #Order 
    sheet.cell(column=1, row=row, value=billOrder[x])
    #Component Name
    sheet.cell(column=2, row=row, value=x)
    #Component Desc
    sheet.cell(column=3, row=row, value=billDesc[x])
    #Amount needed 
    sheet.cell(column=4, row=row, value=bill[x])
    row +=1

wb_obj.save(filename = "CEDIS_MAT.xlsx")
